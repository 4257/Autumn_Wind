import openpyxl

def exlcetest():
    #根据某一列An数据x向下复制对应行数的x*Bn、x*Cb数据至其他行
    # 数据默认从A2开始
    wb = openpyxl.load_workbook("./test.xlsx")
    sheet1 = wb.active

    idlist = []
    isum = 0
    row_num = sheet1.max_row
    # col_num = sheet1.max_column

    for i in range(row_num):
        d1 = sheet1.cell(row=2 + i, column=1).value
        if (d1 != None):
            idlist.append(int(str(d1)))
    for i in range(len(idlist)):
        if (idlist[i] != 0):
            for x in range(idlist[i]):
                sheet1.cell(row=2 + x + isum, column=7).value = sheet1.cell(row=2 + i, column=3).value
                sheet1.cell(row=2 + x + isum, column=8).value = sheet1.cell(row=2 + i, column=4).value
        isum += idlist[i]
    wb.save("./test.xlsx")

if __name__ == '__main__':
    exlcetest()
